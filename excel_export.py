from __future__ import annotations

from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import KotListing

HEADERS = [
    ("Link",               14),
    ("Source",             16),
    ("Title",              40),
    ("Address",            30),
    ("Neighborhood",       16),
    ("Price (€/month)",    15),
    ("Size (m²)",          10),
    ("Price/m²",           10),
    ("Type",               13),
    ("Furnished",          11),
    ("Private bathroom",   16),
    ("Shared bathroom",    16),
    ("Private kitchen",    15),
    ("Shared kitchen",     15),
    ("Internet included",  16),
    ("Utilities included", 16),
    ("Washing machine",    16),
    ("Elevator",           10),
    ("Pets allowed",       12),
    ("Available from",     15),
    ("Scraped on",         17),
]

HEADER_FILL   = PatternFill("solid", fgColor="1F497D")
ROW_FILL_EVEN = PatternFill("solid", fgColor="DCE6F1")

THIN = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)


def _listing_to_row(listing: KotListing) -> list:
    return [
        listing.url,
        listing.source,
        listing.title,
        listing.address,
        listing.neighborhood,
        listing.price_eur_month,
        listing.size_m2,
        listing.price_per_m2,
        listing.listing_type,
        listing.furnished,
        listing.private_bathroom,
        listing.shared_bathroom,
        listing.private_kitchen,
        listing.shared_kitchen,
        listing.internet_included,
        listing.utilities_included,
        listing.washing_machine,
        listing.elevator,
        listing.pets_allowed,
        listing.available_from,
        listing.scraped_on,
    ]


def export_to_excel(listings: List[KotListing], filepath: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Koten Leuven"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(len(listings) + 1, 2)}"

    # ── Header row ──────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (label, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20

    # ── Data rows ───────────────────────────────────────────────────────
    link_font   = Font(color="0563C1", underline="single", size=10)
    normal_font = Font(size=10)

    numeric_cols = {6, 7, 8}  # price, size, price/m²

    for row_idx, listing in enumerate(listings, start=2):
        row_data = _listing_to_row(listing)
        even = row_idx % 2 == 0

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            cell.font = normal_font
            if even:
                cell.fill = ROW_FILL_EVEN

            if col_idx == 1 and value:
                cell.value = "Open listing"
                cell.hyperlink = value
                cell.font = link_font
            elif col_idx in numeric_cols and value:
                try:
                    cell.value = float(value)
                    cell.number_format = "0.0" if col_idx == 8 else "0"
                except ValueError:
                    cell.value = value
            else:
                cell.value = value if value else None  # empty cell when unknown

        ws.row_dimensions[row_idx].height = 16

    # ── Summary sheet ───────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Scraped on"
    ws_summary["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_summary["A2"] = "Listings found"
    ws_summary["B2"] = len(listings)

    sources: dict[str, int] = {}
    for l in listings:
        sources[l.source] = sources.get(l.source, 0) + 1
    ws_summary["A4"] = "Per website"
    for i, (src, count) in enumerate(sources.items(), start=5):
        ws_summary.cell(row=i, column=1, value=src)
        ws_summary.cell(row=i, column=2, value=count)
    ws_summary.column_dimensions["A"].width = 20

    wb.save(filepath)
    print(f"\n  Saved: {filepath}")
    print(f"  Listings: {len(listings)}")
    for src, count in sources.items():
        print(f"    {src}: {count}")
